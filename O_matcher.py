import pandas as pd
import numpy as np
from scipy.stats import norm
import re

import os
import sys
from pathlib import Path

script_path = r"C:\workplace\PythonScripts\O_Matcher"
if script_path not in sys.path:
    sys.path.append(script_path)

from functions import *

from scipy.optimize import brentq

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

####################################

def path_checker(Prompt):
    path = input(Prompt)
    if not Path(path).is_dir():
        print("Directory does not exist. Try again.")
        return path_checker(Prompt)
    else:
        return path


def main():

    out_path = input("Output directory:")
    if not Path(out_path).is_dir():
        os.mkdir(out_path)
    
    output_option = input("Select file to generate:\n1. Intermediate processing file\n2. Final WAGIMS spot-level output\n3. Geoview sample-level output\nYour option: ")
    
    

    UPb_path = path_checker("U-Pb files directory: ")  # path to the U-Pb tables
    O_path = path_checker("Oxygen files directory: ")  # path to the oxygen analyses tables

    UPb = UPb_file_join(UPb_path)
    O_dfs = O_file_join(O_path)

    df_merged = merge_datasets(UPb,O_dfs,"SampleSpot")
    
    df_merged["UPB_ANALYSIS_ID"] = df_merged["UPB_ANALYSIS_ID"].astype("string")

    df_merged = df_merged.drop(["Sample_y"], axis = 1)
    df_merged = df_merged.rename({"Sample_x":"Sample"},axis = 1)

    df_merged = df_merged.sort_values(by = "Unique_O_ID")

    files = dict(zip(df_merged.Unique_O_ID.values,df_merged.file.values))

    col_order = [
    "Session date","Sample", "UPB_ANALYSIS_ID","GroupID","238U_ppm", "232Th_ppm", "Th232U238", "f204_pct",
    "U238Pb206", "U238Pb206_1sig", "Pb207Pb206", "Pb207Pb206_1sig", 
    "4corr_U238Pb206", "4corr_U238Pb206_1sig", "4corr_Pb207Pb206", "4corr_Pb207Pb206_1sig",
    "4corr_86_date","4corr_86_date_1sig","4corr_76_date","4corr_76_date_1sig",
    "Discordance_pct","Material analysed", "Unique_O_ID", "18O/16O", "18O/16O_± rel (%)",
    "16O1H/16O","16O1H/16O_± rel (%)","d18O","± per mil","OH/O","OH/O_± rel (%)","Comment",
    "Laboratory","Instrument"
    ]

    WAGIMS_names = ["Session date","Sample ID","Geochron analysis ID", "Geochron group ID", "238U (ppm)", "232Th (ppm)",
        "232Th/238U", "f204 (%)", "238U/206Pb", "238U/206Pb 1 sigma", "207Pb/206Pb", "207Pb/206Pb 1 sigma",
        "238U/206Pb*", "238U/206Pb* 1 sigma", "207Pb*/206Pb*", "207Pb*/206Pb* 1 sigma",
        "204-corr 206Pb/238U date (Ma)", "204-corr 206Pb/238U date 1 sigma (Ma)",
        "204-corr 207Pb/206Pb date (Ma)", "204-corr 207Pb/206Pb date 1 sigma (Ma)",
        "204-corr 6-38 v 7-6 discordance (%)","Material analysed","Oxygen analysis ID",
        "Measured 18O/16O", "Measured 18O/16O ±1 sigma internal",
        "Measured 16O1H/16O","Measured16O1H/16O ±1 sigma internal (%)",
        "Final δ18O (‰)", "Final δ18O ±1 sigma external (‰)",
        "Drift-corrected 16OH/16O", "Drift-corrected 16OH/16O ±1 sigma internal (%)",
        "Comment","Laboratory","Instrument" 
    ]

    df_merged.insert(len(df_merged.columns), "Material analysed","zircon")
    df_merged.insert(len(df_merged.columns), "Comment","")
    df_merged.insert(len(df_merged.columns), "Laboratory","")
    df_merged.insert(len(df_merged.columns), "Instrument","")
    df_merged.insert(0,"Session date","")


    df_merged_WAGIMS = df_merged[col_order]

    df_merged_WAGIMS = df_merged_WAGIMS.rename(
        dict(zip(col_order,WAGIMS_names), axis = 1)
    )

    Group1 = ["Session date"]
    Group2 = ["Sample ID"]
    Group3 = ["Geochron analysis ID", "Geochron group ID", "238U (ppm)", "232Th (ppm)",
        "232Th/238U", "f204 (%)", "238U/206Pb", "238U/206Pb 1 sigma", "207Pb/206Pb", "207Pb/206Pb 1 sigma",
        "238U/206Pb*", "238U/206Pb* 1 sigma", "207Pb*/206Pb*", "207Pb*/206Pb* 1 sigma",
        "204-corr 206Pb/238U date (Ma)", "204-corr 206Pb/238U date 1 sigma (Ma)",
        "204-corr 207Pb/206Pb date (Ma)", "204-corr 207Pb/206Pb date 1 sigma (Ma)",
        "204-corr 6-38 v 7-6 discordance (%)"]
    Group4 = [
        "Material analysed","Oxygen analysis ID",
        "Measured 18O/16O", "Measured 18O/16O ±1 sigma internal",
        "Measured 16O1H/16O","Measured16O1H/16O ±1 sigma internal (%)"
    ]
    Group5 = [
        "Final δ18O (‰)", "Final δ18O ±1 sigma external (‰)",
        "Drift-corrected 16OH/16O", "Drift-corrected 16OH/16O ±1 sigma internal (%)",
        "Comment"
    ]
    Group6 = [
        "Laboratory","Instrument"
    ]


    Higher_order = [
        "SESSION", "SAMPLE INFORMATION", "GEOCHRONOLOGY DATA", "OXYGEN ANALYSES", "CALCULATED VALUES",
        "LABORATORY INFORMATION"
    ]
    Lower_order = [
        Group1,Group2,Group3,Group4,Group5,Group6
    ]


    MultiIndex = []
    for h,l in zip(Higher_order,Lower_order):
        for c in l:
            MultiIndex.append((h,c))



    # Multi-level column structure
    columns = pd.MultiIndex.from_tuples(
        MultiIndex
    )

    df = pd.DataFrame(df_merged_WAGIMS.values,columns=columns)


    format_1dp = '0.0'
    format_2dp = '0.00'
    format_3dp = '0.000'
    format_4dp = '0.0000' 
    format_5dp = '0.00000'
    format_6dp = '0.000000'
    format_7dp = '0.0000000'

    if output_option == "1":

        df[("FILE","FILE")] = df[("OXYGEN ANALYSES","Oxygen analysis ID")].map(files)

        export_name = input("Intermediate processing table file name (include the .xlsx extension): ")
        
        df.to_excel(os.path.join(out_path,export_name))

        RefMat_df = df[df["SAMPLE INFORMATION"]["Sample ID"] == "GSWA"]
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("257"), "STANDARD ID"] = "M257"
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("tem|tme", case = False), "STANDARD ID"] = "Temora 2"
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("og", case = False), "STANDARD ID"] = "OGC"
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("91500"), "STANDARD ID"] = "91500"
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("cz|CZ", case = False), "STANDARD ID"] = "CZ3"
        RefMat_df.loc[RefMat_df["OXYGEN ANALYSES"]["Oxygen analysis ID"].str.contains("peng", case = False), "STANDARD ID"] = "Penglai"

        RefMat_df = RefMat_df[["SESSION","STANDARD ID","OXYGEN ANALYSES","CALCULATED VALUES","LABORATORY INFORMATION"]]

        Group2 = ["Sample ID",""]
        RefMat_styled = (
            RefMat_df.style
            # ---- LEVEL 0 (top header groups) ----
            .applymap_index(
                lambda v: (
                    'background-color: #FBE5D6; font-weight: bold'
                    if v == 'STANDARD ID' else
                    'background-color: #E2F0D9; font-weight: bold'
                    if v == 'SESSION' else
                    'background-color: #FFFFCC; font-weight: bold'
                    if v == 'OXYGEN ANALYSES' else
                    'background-color: #E2F0D9; font-weight: bold'
                    if v == 'CALCULATED VALUES' else
                    'background-color: #DEEBF7; font-weight: bold'
                    if v == 'LABORATORY INFORMATION' else         
                    ''
                ),
                axis=1,
                level=0
            )

            # ---- LEVEL 1 (sub-columns) ----
            .applymap_index(
                lambda v: (
                    'background-color: #A9D18E; font-weight: bold'
                    if v in Group1 else
                    'background-color: #F8CBAD; font-weight: bold'
                    if v in Group2 else
                    'background-color: #FFE699; font-weight: bold'
                    if v in Group3 else
                    'background-color: #FFFF99; font-weight: bold'
                    if v in Group4 else
                    'background-color: #C5E0B4; font-weight: bold'
                    if v in Group5 else
                    'background-color: #B4C7E7; font-weight: bold'
                    if v in Group6 else
                    ''
                ),
                axis=1,
                level=1
            )
        )

        RefMat_export = input("Reference material table file name: ")
        RefMat_file = os.path.join(out_path,RefMat_export)


        with pd.ExcelWriter(RefMat_file, engine = "openpyxl") as writer:
            RefMat_styled.to_excel(writer, sheet_name = "Oxygen")

            workbook = writer.book
            worksheet = writer.sheets["Oxygen"]

            for cell in worksheet[1]:
                cell.alignment = Alignment(wrap_text = True)
            for cell in worksheet[2]:
                cell.alignment = Alignment(wrap_text = True)

            for i, col in enumerate(RefMat_df.columns, 2):

                if get_column_letter(i) == "F":
                    worksheet.column_dimensions["F"].width = 12
                    for row in range(3, len(RefMat_df) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_7dp
                if get_column_letter(i) == "G":
                    worksheet.column_dimensions["G"].width = 12
                    for row in range(3, len(RefMat_df) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_4dp
                if get_column_letter(i) == "J":
                    worksheet.column_dimensions["J"].width = 12
                    for row in range(3, len(RefMat_df) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp
                if get_column_letter(i) == "K":
                    worksheet.column_dimensions["K"].width = 12
                    for row in range(3, len(RefMat_df) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp


    if output_option == "2":

        df_samples = df.loc[df["SAMPLE INFORMATION"]["Sample ID"] != "GSWA"]

        styled = (
            df_samples.style
            # ---- LEVEL 0 (top header groups) ----
            .applymap_index(
                lambda v: (
                    'background-color: #FBE5D6; font-weight: bold'
                    if v == 'SAMPLE INFORMATION' else
                    'background-color: #FFF2CC; font-weight: bold'
                    if v == 'GEOCHRONOLOGY DATA' else
                    'background-color: #E2F0D9; font-weight: bold'
                    if v == 'SESSION' else
                    'background-color: #FFFFCC; font-weight: bold'
                    if v == 'OXYGEN ANALYSES' else
                    'background-color: #E2F0D9; font-weight: bold'
                    if v == 'CALCULATED VALUES' else
                    'background-color: #DEEBF7; font-weight: bold'
                    if v == 'LABORATORY INFORMATION' else         
                    ''
                ),
                axis=1,
                level=0
            )

            # ---- LEVEL 1 (sub-columns) ----
            .applymap_index(
                lambda v: (
                    'background-color: #A9D18E; font-weight: bold'
                    if v in Group1 else
                    'background-color: #F8CBAD; font-weight: bold'
                    if v in Group2 else
                    'background-color: #FFE699; font-weight: bold'
                    if v in Group3 else
                    'background-color: #FFFF99; font-weight: bold'
                    if v in Group4 else
                    'background-color: #C5E0B4; font-weight: bold'
                    if v in Group5 else
                    'background-color: #B4C7E7; font-weight: bold'
                    if v in Group6 else
                    ''
                ),
                axis=1,
                level=1
            )
        )




        export_name = input("export file name (include the .xlsx extension): ")
        export_name = os.path.join(
            out_path,export_name
        )

        with pd.ExcelWriter(export_name, engine="openpyxl") as writer:
            styled.to_excel(writer, sheet_name="Oxygen")

            workbook  = writer.book
            worksheet = writer.sheets["Oxygen"]

            ######### row heights

            for cell in worksheet[1]:
                cell.alignment = Alignment(wrap_text=True)
            for cell in worksheet[2]:
                cell.alignment = Alignment(wrap_text=True)


            ######### column widths
            for i, col in enumerate(df_samples.columns, 2):

                if get_column_letter(i) == "B":
                    worksheet.column_dimensions["B"].width = 10.3
                if get_column_letter(i) == "C":
                    worksheet.column_dimensions["C"].width = 15
                if get_column_letter(i) == "D":
                    worksheet.column_dimensions["D"].width = 10.71
                if get_column_letter(i) == "E":
                    worksheet.column_dimensions["E"].width = 9
                if get_column_letter(i) == "F":
                    worksheet.column_dimensions["F"].width = 6.6
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "G":
                    worksheet.column_dimensions["G"].width = 6.6
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "H":
                    worksheet.column_dimensions["H"].width = 6.6
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp
                if get_column_letter(i) == "I":
                    worksheet.column_dimensions["I"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "I":
                    worksheet.column_dimensions["I"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "J":
                    worksheet.column_dimensions["J"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "K":
                    worksheet.column_dimensions["K"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "L":
                    worksheet.column_dimensions["L"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_5dp
                if get_column_letter(i) == "M":
                    worksheet.column_dimensions["M"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_5dp
                if get_column_letter(i) == "N":
                    worksheet.column_dimensions["N"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "O":
                    worksheet.column_dimensions["O"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "P":
                    worksheet.column_dimensions["P"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_5dp
                if get_column_letter(i) == "Q":
                    worksheet.column_dimensions["Q"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_5dp
                if get_column_letter(i) == "R":
                    worksheet.column_dimensions["R"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "S":
                    worksheet.column_dimensions["S"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "T":
                    worksheet.column_dimensions["T"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "U":
                    worksheet.column_dimensions["U"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = "0"
                if get_column_letter(i) == "V":
                    worksheet.column_dimensions["V"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp
                if get_column_letter(i) == "X":
                    worksheet.column_dimensions["X"].width = 13
                if get_column_letter(i) == "Y":
                    worksheet.column_dimensions["Y"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_7dp
                if get_column_letter(i) == "Z":
                    worksheet.column_dimensions["Z"].width = 8.5
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_5dp
                if get_column_letter(i) == "AA":
                    worksheet.column_dimensions["AA"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_6dp
                if get_column_letter(i) == "AB":
                    worksheet.column_dimensions["AB"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_6dp
                if get_column_letter(i) == "AC":
                    worksheet.column_dimensions["AC"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp
                if get_column_letter(i) == "AD":
                    worksheet.column_dimensions["AD"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_2dp
                if get_column_letter(i) == "AE":
                    worksheet.column_dimensions["AE"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "AF":
                    worksheet.column_dimensions["AF"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_3dp
                if get_column_letter(i) == "AG":
                    worksheet.column_dimensions["AG"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_4dp
                if get_column_letter(i) == "AH":
                    worksheet.column_dimensions["AH"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_4dp
                if get_column_letter(i) == "AJ":
                    worksheet.column_dimensions["AJ"].width = 12
                    for row in range(3, len(df_samples) + 3):
                        worksheet.cell(row=row,column=i).number_format = format_6dp

    if output_option == "3":
        sample_level_name = input("Sample-level file name (include extension):")
        sample_level = os.path.join(out_path,sample_level_name)

        df_merged_filtered = df_merged.loc[df_merged.GroupID == "I"]

        sample_weighted_mean = (
            df_merged_filtered
            .groupby("Sample")
            .apply(lambda g: weighted_mean(
                g['4corr_Pb207Pb206'],
                g['4corr_Pb207Pb206_1sig'],
                method = "internal_sigma"
            ))
        )


        sample_weighted_mean = pd.DataFrame.from_dict(sample_weighted_mean.to_dict(), orient="index")

        sample_weighted_mean[["age", "sigma_age"]] = (
            sample_weighted_mean.apply(
                lambda s: pb207_pb206_age_with_uncertainty(
                    s["mean"], 2*s["uncertainty"]
                ),
                axis=1,
                result_type="expand"
            )
        )

        df_merged.loc[df_merged.GroupID.isin(["I"]),"Igneous_age"] = \
            df_merged.loc[df_merged.GroupID.isin(["I"]),"Sample"].map(
                sample_weighted_mean["age"]
            )

        df_merged.loc[df_merged.GroupID.isin(["I"]),"Igneous_age_unc"] = \
            df_merged.loc[df_merged.GroupID.isin(["I"]),"Sample"].map(
                sample_weighted_mean["sigma_age"]
            )

        df_merged.loc[~df_merged.GroupID.isin(["I"]),"Igneous_age"] = \
            df_merged.loc[~df_merged.GroupID.isin(["I"]),"4corr_76_date"]

        df_merged.loc[~df_merged.GroupID.isin(["I"]),"Igneous_age_unc"] = \
            df_merged.loc[~df_merged.GroupID.isin(["I"]),"4corr_76_date_1sig"]

        Age = df_merged.loc[df_merged.GroupID.isin(["I"])].groupby("Sample")["Igneous_age"].first().round(1)
        Age_unc = df_merged.loc[df_merged.GroupID.isin(["I"])].groupby("Sample")["Igneous_age_unc"].first().round(1)

        variables = ["d18O"]
        uncertainties = ["± per mil"]
        prepends = ["D18O"]

        composite = create_aggregate_df(df_merged_filtered,"Sample",variables,uncertainties,prepends,by_list="exclude.txt")

        composite["AGE"] = composite.index.map(Age)
        composite["AGE_UNC"] = composite.index.map(Age_unc)

        composite.to_excel(sample_level)

    elif output_option not in ["1","2","3"]: 
        print("Invalid option. Closing application.")

    
if __name__ == "__main__":
    main()



        


